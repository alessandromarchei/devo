import os
import sys
import h5py
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def get_folder_size(path):
    total = 0
    for dirpath, _, filenames in os.walk(path):
        for f in filenames:
            try:
                total += os.path.getsize(os.path.join(dirpath, f))
            except FileNotFoundError:
                continue
    return total

def get_file_size(path):
    try:
        return os.path.getsize(path)
    except FileNotFoundError:
        return 0

def analyze_scene(scene_path):
    parts = scene_path.strip().split('/')
    try:
        scene_name = parts[-3]
        print(f"\nAnalyzing scene: {scene_name}")
        difficulty = parts[-2]
        print(f"  - Difficulty: {difficulty}")
        scene_id = parts[-1]
        print(f"  - Scene ID: {scene_id}")
    except IndexError:
        scene_name, difficulty, scene_id = 'unknown', 'unknown', 'unknown'

    folder_size_gb = get_folder_size(scene_path) / (1024 ** 3)
    events_file = os.path.join(scene_path, "events.h5")
    events_size_gb = get_file_size(events_file) / (1024 ** 3)
    print(f"  - Folder size: {folder_size_gb:.2f} GB")

    num_events = -1
    if os.path.exists(events_file):
        try:
            with h5py.File(events_file, 'r') as f:
                num_events = len(f["events"]["x"])
                print(f"  - Number of events: {num_events}")
                num_events = num_events / 1e9  # Convert to billions of events
        except Exception:
            pass

    timestamps_file = os.path.join(scene_path, "timestamps.txt")
    depth_folder = os.path.join(scene_path, "depth_left")

    #count number of files inside depth_left folder
    if os.path.exists(depth_folder):
        try:
            num_frames = len([f for f in os.listdir(depth_folder) if f.endswith('.npy')])
            print(f"  - Number of frames in depth_left: {num_frames}")
        except Exception:
            num_frames = -1
    else:
        num_frames = -1
    num_timestamps, duration = -1, -1


    if os.path.exists(timestamps_file):
        try:
            with open(timestamps_file, 'r') as f:
                lines = f.readlines()
                num_timestamps = len(lines)
                if num_timestamps > 1:
                    print(f"  - Number of timestamps: {num_timestamps}")

        except Exception:
            pass
    
    duration = num_frames * 0.03
    print(f"  - Scene duration: {duration:.2f} seconds")

    return [
        scene_name,
        difficulty,
        scene_id,
        folder_size_gb,
        events_size_gb,
        num_events,
        num_timestamps,
        num_frames,
        duration
    ]

def merge_column_cells(ws, col_idx, values, start_row=2):
    current_value = None
    merge_start = start_row

    for i, value in enumerate(values, start=start_row):
        if value != current_value:
            if i - merge_start > 1:
                ws.merge_cells(
                    start_row=merge_start, start_column=col_idx,
                    end_row=i - 1, end_column=col_idx
                )
                ws.cell(merge_start, col_idx).alignment = Alignment(vertical="center", horizontal="center")
            current_value = value
            merge_start = i
        ws.cell(i, col_idx, value)

    # Final group
    if len(values) + start_row - merge_start > 1:
        ws.merge_cells(
            start_row=merge_start, start_column=col_idx,
            end_row=start_row + len(values) - 1, end_column=col_idx
        )
        ws.cell(merge_start, col_idx).alignment = Alignment(vertical="center", horizontal="center")
from openpyxl.styles import PatternFill

def write_grouped_excel(rows, output_path="output.xlsx"):
    from operator import itemgetter
    wb = Workbook()
    ws = wb.active
    ws.title = "Scene Analysis"

    headers = [
        "Scene Name", "Difficulty", "Scene ID",
        "Folder Size (GB)", "events.h5 Size (GB)",
        "Number of Events", "Number of Timestamps",
        "Number of Frames", "Duration (s)"
    ]
    ws.append(headers)

    # Sort rows by scene name then difficulty
    rows_sorted = sorted(rows, key=itemgetter(0, 1))

    # For coloring
    pastel_fills = [
        "FFFDE0", "E0F7FA", "E8F5E9", "F3E5F5", "FBE9E7",
        "FFF3E0", "E1F5FE", "F1F8E9", "FCE4EC", "E8EAF6"
    ]
    color_index = 0

    scene_names = []
    difficulties = []
    current_scene = rows_sorted[0][0]
    group_start_row = 2
    current_fill = PatternFill(start_color=pastel_fills[color_index],
                               end_color=pastel_fills[color_index], fill_type="solid")

    for i, row in enumerate(rows_sorted):
        scene_names.append(row[0])
        difficulties.append(row[1])
        ws.append(['', '', *row[2:]])

        # If next row is a different scene, apply color fill to group
        is_last = i == len(rows_sorted) - 1
        next_scene = rows_sorted[i + 1][0] if not is_last else None

        if row[0] != next_scene:
            for r in range(group_start_row, group_start_row + (i - (group_start_row - 2))):
                for c in range(1, len(headers)+1):
                    ws.cell(r, c).fill = current_fill
            # Prepare next group
            group_start_row = i + 3
            color_index = (color_index + 1) % len(pastel_fills)
            current_fill = PatternFill(start_color=pastel_fills[color_index],
                                       end_color=pastel_fills[color_index], fill_type="solid")

    # Merge Scene Name and Difficulty columns
    merge_column_cells(ws, 1, scene_names)
    merge_column_cells(ws, 2, difficulties)

    wb.save(output_path)
    print(f"✅ Excel file saved as {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python analyze_tartanevent_excel_grouped_sorted.py paths.txt")
        sys.exit(1)

    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        print(f"❌ Error: {input_file} not found.")
        sys.exit(1)

    with open(input_file, 'r') as f:
        scene_paths = [line.strip() for line in f if line.strip()]

    all_rows = [analyze_scene(path) for path in scene_paths]
    write_grouped_excel(all_rows)
